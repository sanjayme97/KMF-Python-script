import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Replace 'file_path' with the actual file path of your Excel file
file_path = 'C:\\Users\\sanjay.me\\Downloads\\2023 nov 1.xls'

# Read the Excel file and specify data types for columns
data = pd.read_excel(file_path, usecols=[1, 2, 3, 10], dtype={2: str})
data[data.columns[3]] = data[data.columns[3]].apply(np.floor)
filtered_data = data[data[data.columns[3]] > 0]
# Define the 'Bank Name' column with a default value
bank_name = 'HDCC Hemavathi Branch Hsn'
filtered_data.insert(3, 'Bank Name', bank_name) 
filtered_data = filtered_data.reset_index(drop=True)
 # Insert the 'Bank Name' column after 'Account Number'
filtered_data.insert(0, 'S No.', filtered_data.index+1)  # Insert the index as the first column
print(filtered_data)
# Save the modified data as a new Excel file
# Replace 'new_modified_file.xlsx' with your desired file name for the output
# filtered_data.to_excel('C:\\Users\\sanjay.me\\Downloads\\new_modified_file.xlsx', index=False)

# Load the saved Excel file for adjusting column widths and creating a table
# workbook = openpyxl.load_workbook('C:\\Users\\sanjay.me\\Downloads\\December Payment Report-23.xlsx')
workbook = Workbook()
sheet = workbook.active
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set a value in cell A1
# Merge cells and set the alignment to center
merged_cell = sheet['A1']
merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

header_list = filtered_data.columns.tolist()
print(header_list)
for idx, header in enumerate(header_list, start=1):
    sheet.cell(row=4, column=idx).value = header

for r_idx, row in filtered_data.iterrows():
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx + 5, column=c_idx).value = value

# Convert the data into an Excel Table (ListObject)
table = openpyxl.worksheet.table.Table(displayName="MyTable", ref=f"A4:{openpyxl.utils.get_column_letter(filtered_data.shape[1])}{filtered_data.shape[0] + 4}")
table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
sheet.add_table(table)

# Adjust column widths
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
# Save the Excel file with adjusted column widths and the table
sheet['A1'] = ' Kondajji Koppalu Oct-2023 Payment'
sheet.merge_cells('A1:F3')

last_column_sum = filtered_data.iloc[:, -1].sum()

# Get the last row index to insert the sum
last_row = filtered_data.shape[0] + 5  # Offset by the header and table start rows

# Insert the sum at the end of the last column in the table
sheet.cell(row=last_row, column=filtered_data.shape[1]).value = last_column_sum
workbook.save('C:\\Users\\sanjay.me\\Downloads\\PaymentReport.xlsx')
